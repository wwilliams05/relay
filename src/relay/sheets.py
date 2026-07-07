"""Tracker adapter (PRD §4).

Two interchangeable backends behind the `Tracker` Protocol:

- `LocalXlsxTracker` — an openpyxl workbook on disk. Default; needs no credentials,
  so the whole N0–N6 funnel is runnable and testable today.
- `SheetsTracker`  — Google Sheets via gspread (drop-in, same interface).

Writes are **upserts**: re-running discovery refreshes the machine-owned columns
(title, email, hook, …) while preserving the human-owned gate columns you've edited
in the tracker (`want_to_message`, `referral_cleared`, funnel state, …).
"""

from __future__ import annotations

from datetime import date
from pathlib import Path
from typing import Any, Protocol

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .models import Contact, EmailStatus, Job, Project, Target, Why

# URL columns rendered as clickable hyperlinks in the workbook.
URL_COLUMNS = {"job_url", "profile_url", "jd_url"}
_HYPERLINK_FONT = Font(color="0563C1", underline="single")

# Boolean gate columns: rendered as native Excel checkboxes (see xlsx_checkbox.py).
# Cap column width so long free-text columns (hook, chat_notes, prd_prompt) stay
# readable rather than sprawling.
BOOL_COLUMNS = {
    "want_to_message", "referral_cleared", "draft_created", "responded",
    "interested", "pursue",
}
MAX_COL_WIDTH = 80


class Tracker(Protocol):
    """The interface every storage backend implements (Sheets or local xlsx)."""

    def upsert_target(self, target: Target) -> None: ...
    def write_contacts(self, contacts: list[Contact]) -> None: ...
    def read_contacts(self) -> list[Contact]: ...
    def update_contact(self, contact: Contact) -> None: ...
    def write_projects(self, projects: list[Project]) -> None: ...
    def read_projects(self) -> list[Project]: ...
    def write_jobs(self, jobs: list[Job]) -> None: ...
    def read_jobs(self) -> list[Job]: ...


# --- Tab schemas: column header -> how to (de)serialize a model field ---------
#
# Each tab is described once, as an ordered list of (header, getter, setter) so the
# workbook layout, the human-facing headers, and the round-trip logic stay in sync.

TARGET_COLUMNS = ["company", "role", "jd_url", "anchor_framing", "status", "similar_titles"]
CONTACT_COLUMNS = [
    "name", "title", "company", "profile_url", "why", "school_match",
    "email", "email_status", "hook", "want_to_message", "referral_cleared",
    "draft_created", "messaged_date", "responded", "chat_notes", "next_step",
]
PROJECT_COLUMNS = [
    "target_company", "for_contact", "project_idea", "skills_shown",
    "interested", "prd_prompt",
]
JOB_COLUMNS = [
    "company", "title", "location", "source", "job_url",
    "date_posted", "fit_score", "fit_reason", "pursue", "status",
]

# Columns a human owns in the tracker; discovery must not clobber these on re-run.
CONTACT_HUMAN_COLUMNS = {
    "want_to_message", "referral_cleared", "draft_created",
    "messaged_date", "responded", "chat_notes", "next_step",
}


def _autofit(ws, headers: list[str]) -> None:
    """Size each column to its widest cell (header or data), capped for readability."""
    for idx, header in enumerate(headers, start=1):
        width = len(str(header))
        for row in range(2, ws.max_row + 1):
            value = ws.cell(row=row, column=idx).value
            if value is not None:
                width = max(width, len(str(value)))
        ws.column_dimensions[get_column_letter(idx)].width = min(width + 2, MAX_COL_WIDTH)


def _cell_for(header: str, value: Any) -> Any:
    """Boolean gate columns become real booleans (plain FALSE/TRUE); rest stringify."""
    if header in BOOL_COLUMNS:
        return _as_bool(value)
    return _to_cell(value)


def _to_cell(value: Any) -> Any:
    """Model value -> a plain, spreadsheet-friendly cell value."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, list):
        return ", ".join(str(v) for v in value)
    if isinstance(value, (Why, EmailStatus)):
        return value.value
    return value


def _cell_str(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _as_bool(value: Any) -> bool:
    return _cell_str(value).upper() in {"TRUE", "1", "YES", "X", "☑", "✓"}


def _as_list(value: Any) -> list[str]:
    s = _cell_str(value)
    return [part.strip() for part in s.split(",") if part.strip()]


def _as_date(value: Any) -> date | None:
    s = _cell_str(value)
    if not s:
        return None
    if isinstance(value, date):
        return value
    return date.fromisoformat(s[:10])


def target_to_row(t: Target) -> dict[str, Any]:
    return {
        "company": t.company, "role": t.role, "jd_url": _to_cell(t.jd_url),
        "anchor_framing": t.anchor_framing, "status": t.status,
        "similar_titles": _to_cell(t.similar_titles),
    }


def contact_to_row(c: Contact) -> dict[str, Any]:
    return {
        "name": c.name, "title": _to_cell(c.title), "company": c.company,
        "profile_url": _to_cell(c.profile_url), "why": _to_cell(c.why),
        "school_match": _to_cell(c.school_match), "email": _to_cell(c.email),
        "email_status": _to_cell(c.email_status), "hook": _to_cell(c.hook),
        "want_to_message": _to_cell(c.want_to_message),
        "referral_cleared": _to_cell(c.referral_cleared),
        "draft_created": _to_cell(c.draft_created),
        "messaged_date": _to_cell(c.messaged_date),
        "responded": _to_cell(c.responded),
        "chat_notes": _to_cell(c.chat_notes), "next_step": _to_cell(c.next_step),
    }


def row_to_contact(row: dict[str, Any]) -> Contact:
    return Contact(
        name=_cell_str(row.get("name")),
        title=_cell_str(row.get("title")) or None,
        company=_cell_str(row.get("company")),
        profile_url=_cell_str(row.get("profile_url")) or None,
        why=Why(_cell_str(row.get("why")) or "similar_role"),
        school_match=_cell_str(row.get("school_match")) or None,
        email=_cell_str(row.get("email")) or None,
        email_status=EmailStatus(_cell_str(row.get("email_status")) or "unavailable"),
        hook=_cell_str(row.get("hook")) or None,
        want_to_message=_as_bool(row.get("want_to_message")),
        referral_cleared=_as_bool(row.get("referral_cleared")),
        draft_created=_as_bool(row.get("draft_created")),
        messaged_date=_as_date(row.get("messaged_date")),
        responded=_as_bool(row.get("responded")),
        chat_notes=_cell_str(row.get("chat_notes")) or None,
        next_step=_cell_str(row.get("next_step")) or None,
    )


def project_to_row(p: Project) -> dict[str, Any]:
    return {
        "target_company": p.target_company, "for_contact": _to_cell(p.for_contact),
        "project_idea": p.project_idea, "skills_shown": _to_cell(p.skills_shown),
        "interested": _to_cell(p.interested), "prd_prompt": _to_cell(p.prd_prompt),
    }


def row_to_project(row: dict[str, Any]) -> Project:
    return Project(
        target_company=_cell_str(row.get("target_company")),
        for_contact=_cell_str(row.get("for_contact")) or None,
        project_idea=_cell_str(row.get("project_idea")),
        skills_shown=_as_list(row.get("skills_shown")),
        interested=_as_bool(row.get("interested")),
        prd_prompt=_cell_str(row.get("prd_prompt")) or None,
    )


def job_to_row(j: Job) -> dict[str, Any]:
    return {
        "company": j.company, "title": j.title, "location": _to_cell(j.location),
        "job_type": _to_cell(j.job_type), "source": _to_cell(j.source),
        "job_url": _to_cell(j.job_url), "date_posted": _to_cell(j.date_posted),
        "fit_score": j.fit_score, "fit_reason": _to_cell(j.fit_reason),
        "pursue": _to_cell(j.pursue), "status": j.status,
    }


def row_to_job(row: dict[str, Any]) -> Job:
    raw_score = _cell_str(row.get("fit_score"))
    return Job(
        company=_cell_str(row.get("company")),
        title=_cell_str(row.get("title")),
        location=_cell_str(row.get("location")) or None,
        job_type=_cell_str(row.get("job_type")) or None,
        source=_cell_str(row.get("source")) or None,
        job_url=_cell_str(row.get("job_url")) or None,
        date_posted=_as_date(row.get("date_posted")),
        fit_score=int(float(raw_score)) if raw_score else 0,
        fit_reason=_cell_str(row.get("fit_reason")) or None,
        pursue=_as_bool(row.get("pursue")),
        status=_cell_str(row.get("status")) or "new",
    )


def job_key(j: Job | dict[str, Any]) -> str:
    """Stable identity for a job: prefer URL, fall back to company+title."""
    get = j.get if isinstance(j, dict) else lambda k: getattr(j, k, None)
    url = _cell_str(get("job_url"))
    if url:
        return f"url::{url.lower()}"
    return f"ct::{_cell_str(get('company')).lower()}::{_cell_str(get('title')).lower()}"


def contact_key(c: Contact | dict[str, Any]) -> str:
    """Stable identity for a contact: prefer email, fall back to name+company."""
    get = c.get if isinstance(c, dict) else lambda k: getattr(c, k, None)
    email = _cell_str(get("email"))
    if email:
        return f"email::{email.lower()}"
    return f"nc::{_cell_str(get('name')).lower()}::{_cell_str(get('company')).lower()}"


def project_key(p: Project | dict[str, Any]) -> str:
    """Stable identity for a project idea: company + the idea text."""
    get = p.get if isinstance(p, dict) else lambda k: getattr(p, k, None)
    return (f"{_cell_str(get('target_company')).lower()}"
            f"::{_cell_str(get('project_idea')).lower()}")


# --- backend-agnostic merge semantics ------------------------------------------
#
# Both trackers read whatever rows the backend holds, merge new model rows into
# them with these pure functions, and write the result back — so the upsert rules
# (human-owned columns survive, jobs stay sorted/pruned) live in exactly one place.

def merge_target_rows(existing: list[dict[str, Any]], target: Target) -> list[dict[str, Any]]:
    key = (target.company.lower(), target.role.lower())
    rows = [r for r in existing
            if (_cell_str(r.get("company")).lower(), _cell_str(r.get("role")).lower()) != key]
    rows.append(target_to_row(target))
    return rows


def merge_contact_rows(existing: list[dict[str, Any]],
                       contacts: list[Contact]) -> list[dict[str, Any]]:
    """Upsert by contact_key, preserving human-owned columns on existing rows."""
    by_key = {contact_key(r): r for r in existing}
    merged: dict[str, dict[str, Any]] = dict(by_key)
    for c in contacts:
        k = contact_key(c)
        row = contact_to_row(c)
        if k in by_key:
            for col in CONTACT_HUMAN_COLUMNS:
                row[col] = by_key[k].get(col, row[col])
        merged[k] = row
    return list(merged.values())


def replace_contact_row(existing: list[dict[str, Any]],
                        contact: Contact) -> list[dict[str, Any]]:
    """Overwrite a single contact's row wholesale (used by N5/N6)."""
    k = contact_key(contact)
    rows = [r for r in existing if contact_key(r) != k]
    rows.append(contact_to_row(contact))
    return rows


def merge_project_rows(existing: list[dict[str, Any]],
                       projects: list[Project]) -> list[dict[str, Any]]:
    """Upsert by project_key: the human `interested` box survives, and an existing
    prd_prompt is never clobbered by an empty one."""
    by_key = {project_key(r): r for r in existing}
    merged: dict[str, dict[str, Any]] = dict(by_key)
    for p in projects:
        k = project_key(p)
        row = project_to_row(p)
        if k in by_key:
            row["interested"] = by_key[k].get("interested", row["interested"])
            if not _cell_str(row.get("prd_prompt")):
                row["prd_prompt"] = by_key[k].get("prd_prompt", "")
        merged[k] = row
    return list(merged.values())


def _is_sample_job_row(row: dict[str, Any]) -> bool:
    """Demo/fixture data that must never persist alongside real postings: rows the
    fixture adapter wrote (source=fixture) or anything pointing at example.com (the
    fixture URLs — also catches rows written before fixtures were labeled)."""
    if _cell_str(row.get("source")).lower() == "fixture":
        return True
    return "example.com" in _cell_str(row.get("job_url")).lower()


def merge_job_rows(existing: list[dict[str, Any]], jobs: list[Job],
                   floor: int) -> list[dict[str, Any]]:
    """Upsert by job_key, keep the human `pursue`/status, prune sub-floor rows that
    aren't pursued, and order highest fit first."""
    by_key = {job_key(r): r for r in existing}
    merged: dict[str, dict[str, Any]] = dict(by_key)
    incoming_real = False
    for j in jobs:
        k = job_key(j)
        row = job_to_row(j)
        if not _is_sample_job_row(row):
            incoming_real = True
        if k in by_key:
            row["pursue"] = by_key[k].get("pursue", row["pursue"])
            row["status"] = by_key[k].get("status", row["status"])
        merged[k] = row

    # Real postings evict demo rows — even pursued ones; a ticked fake posting is
    # still fake. An all-sample batch (explicit fixture demo mode) keeps them so the
    # offline demo flow works.
    if incoming_real:
        merged = {k: r for k, r in merged.items() if not _is_sample_job_row(r)}

    def _fit(row: dict[str, Any]) -> int:
        try:
            return int(float(row.get("fit_score") or 0))
        except (TypeError, ValueError):
            return 0

    def _checked(row: dict[str, Any]) -> bool:
        p = row.get("pursue")
        return p is True or _cell_str(p).lower() in {"true", "1", "yes"}

    kept = [r for r in merged.values() if _fit(r) >= floor or _checked(r)]
    return sorted(kept, key=_fit, reverse=True)


class LocalXlsxTracker:
    """openpyxl-backed tracker. One workbook, one sheet per model tab."""

    def __init__(self, path: str | Path) -> None:
        self.path = Path(path)

    # -- workbook plumbing ----------------------------------------------------
    def _save(self, wb: Workbook) -> None:
        """Save the workbook, then upgrade boolean cells to Excel checkboxes.

        Injection is best-effort: if the workbook XML isn't shaped as expected we keep
        the plain-boolean file rather than risk shipping one Excel wants to repair.
        """
        from . import config
        from .xlsx_checkbox import CheckboxInjectionError, inject_checkboxes

        try:
            wb.save(self.path)
        except PermissionError as exc:
            raise RuntimeError(
                f"Couldn't save {self.path.name} — it's open in Excel (or locked). "
                "Close the spreadsheet, then run this step again."
            ) from exc
        if config.xlsx_checkboxes():
            try:
                inject_checkboxes(self.path)
            except (CheckboxInjectionError, PermissionError, OSError):
                pass  # leave the plain FALSE/TRUE workbook in place

    def _load(self) -> Workbook:
        if self.path.exists():
            return load_workbook(self.path)
        return Workbook()  # a fresh book with a default "Sheet"

    def _sheet(self, wb: Workbook, title: str, headers: list[str]):
        # Always create the tab fresh with the header as row 1. (Reusing openpyxl's
        # default "Sheet" bumped the header to row 2 via a phantom empty row 1.)
        ws = wb.create_sheet(title)
        ws.append(headers)
        # Drop the stray default sheet a brand-new workbook ships with.
        if "Sheet" in wb.sheetnames and wb["Sheet"] is not ws:
            stray = wb["Sheet"]
            if stray.max_row <= 1 and stray.max_column <= 1:
                del wb["Sheet"]
        return ws

    def _read_rows(self, title: str, headers: list[str]) -> list[dict[str, Any]]:
        if not self.path.exists():
            return []
        # read_only holds the file open until close(); on Windows a lingering handle
        # can block the next save, so always close it (try/finally around the read).
        wb = load_workbook(self.path, read_only=True)
        try:
            if title not in wb.sheetnames:
                return []
            ws = wb[title]
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header_row = list(next(rows_iter))
            except StopIteration:
                return []
            header_row = [_cell_str(h) for h in header_row]
            out: list[dict[str, Any]] = []
            for raw in rows_iter:
                if raw is None or all(c is None or _cell_str(c) == "" for c in raw):
                    continue
                out.append({h: raw[i] if i < len(raw) else None for i, h in enumerate(header_row)})
            return out
        finally:
            wb.close()

    def _write_rows(self, wb: Workbook, title: str, headers: list[str],
                    rows: list[dict[str, Any]]) -> None:
        # Rebuild the sheet from scratch so column order/headers stay authoritative.
        if title in wb.sheetnames:
            del wb[title]
        ws = self._sheet(wb, title, headers)
        url_cols = {h: headers.index(h) + 1 for h in headers if h in URL_COLUMNS}
        for r_idx, row in enumerate(rows, start=2):  # header occupies row 1
            ws.append([_cell_for(h, row.get(h, "")) for h in headers])
            for h, col in url_cols.items():  # make URL cells clickable
                url = _cell_str(row.get(h, ""))
                if url.startswith("http"):
                    cell = ws.cell(row=r_idx, column=col)
                    cell.hyperlink = url
                    cell.font = _HYPERLINK_FONT
        _autofit(ws, headers)

    # -- Targets --------------------------------------------------------------
    def upsert_target(self, target: Target) -> None:
        rows = merge_target_rows(self._read_rows("Targets", TARGET_COLUMNS), target)
        wb = self._load()
        self._write_rows(wb, "Targets", TARGET_COLUMNS, rows)
        self._save(wb)

    # -- Contacts -------------------------------------------------------------
    def write_contacts(self, contacts: list[Contact]) -> None:
        rows = merge_contact_rows(self._read_rows("Contacts", CONTACT_COLUMNS), contacts)
        wb = self._load()
        self._write_rows(wb, "Contacts", CONTACT_COLUMNS, rows)
        self._save(wb)

    def read_contacts(self) -> list[Contact]:
        return [row_to_contact(r) for r in self._read_rows("Contacts", CONTACT_COLUMNS)]

    def update_contact(self, contact: Contact) -> None:
        rows = replace_contact_row(self._read_rows("Contacts", CONTACT_COLUMNS), contact)
        wb = self._load()
        self._write_rows(wb, "Contacts", CONTACT_COLUMNS, rows)
        self._save(wb)

    # -- Projects -------------------------------------------------------------
    def write_projects(self, projects: list[Project]) -> None:
        rows = merge_project_rows(self._read_rows("Projects", PROJECT_COLUMNS), projects)
        wb = self._load()
        self._write_rows(wb, "Projects", PROJECT_COLUMNS, rows)
        self._save(wb)

    def read_projects(self) -> list[Project]:
        return [row_to_project(r) for r in self._read_rows("Projects", PROJECT_COLUMNS)]

    # -- Jobs -----------------------------------------------------------------
    def write_jobs(self, jobs: list[Job]) -> None:
        from . import config

        rows = merge_job_rows(self._read_rows("Jobs", JOB_COLUMNS), jobs,
                              floor=config.jobs_min_fit())
        wb = self._load()
        self._write_rows(wb, "Jobs", JOB_COLUMNS, rows)
        self._save(wb)

    def read_jobs(self) -> list[Job]:
        return [row_to_job(r) for r in self._read_rows("Jobs", JOB_COLUMNS)]


def _gspread_client():
    """A gspread client authed with the service account from .env. Import + auth are
    deferred to first use so the default xlsx path never needs Google installed."""
    from . import config

    try:
        import gspread  # already a project dep, but keep the error actionable
    except ImportError as exc:
        raise RuntimeError("gspread not installed — pip install gspread google-auth") from exc
    creds = config.google_credentials_path()
    if not creds.exists():
        raise RuntimeError(
            f"Google service-account json not found at {creds} — set "
            "GOOGLE_APPLICATION_CREDENTIALS (see .env.example) and share the "
            "spreadsheet with the service account's email."
        )
    return gspread.service_account(filename=str(creds))


def _worksheet_missing_excs() -> tuple[type[Exception], ...]:
    """gspread's WorksheetNotFound when available; KeyError covers injected fakes."""
    try:
        from gspread.exceptions import WorksheetNotFound

        return (WorksheetNotFound, KeyError)
    except ImportError:
        return (KeyError,)


class SheetsTracker:
    """Google Sheets implementation — same merge semantics as LocalXlsxTracker,
    via the shared merge_* functions.

    `client` is injectable (anything with .open_by_key) so tests run against an
    in-memory fake; the real path lazily builds a gspread service-account client.
    Sheets returns every cell as a string, which the row converters already parse
    (booleans arrive as "TRUE"/"FALSE" — same shape the xlsx path round-trips).
    """

    def __init__(self, workbook_key: str, client: Any | None = None) -> None:
        self.workbook_key = workbook_key
        self._client = client

    # -- worksheet plumbing ----------------------------------------------------
    def _spreadsheet(self):
        if self._client is None:
            self._client = _gspread_client()
        return self._client.open_by_key(self.workbook_key)

    def _worksheet(self, title: str, create: bool):
        ss = self._spreadsheet()
        try:
            return ss.worksheet(title)
        except _worksheet_missing_excs():
            if not create:
                return None
            return ss.add_worksheet(title=title, rows=200, cols=26)

    def _read_rows(self, title: str) -> list[dict[str, Any]]:
        ws = self._worksheet(title, create=False)
        if ws is None:
            return []
        values = ws.get_all_values()
        if not values:
            return []
        headers = [_cell_str(h) for h in values[0]]
        out: list[dict[str, Any]] = []
        for raw in values[1:]:
            if all(_cell_str(c) == "" for c in raw):
                continue
            out.append({h: (raw[i] if i < len(raw) else "") for i, h in enumerate(headers)})
        return out

    def _write_rows(self, title: str, headers: list[str],
                    rows: list[dict[str, Any]]) -> None:
        ws = self._worksheet(title, create=True)
        data = [headers] + [[_cell_for(h, r.get(h, "")) for h in headers] for r in rows]
        ws.clear()
        ws.update(values=data, range_name="A1")

    # -- Tracker protocol --------------------------------------------------------
    def upsert_target(self, target: Target) -> None:
        self._write_rows("Targets", TARGET_COLUMNS,
                         merge_target_rows(self._read_rows("Targets"), target))

    def write_contacts(self, contacts: list[Contact]) -> None:
        self._write_rows("Contacts", CONTACT_COLUMNS,
                         merge_contact_rows(self._read_rows("Contacts"), contacts))

    def read_contacts(self) -> list[Contact]:
        return [row_to_contact(r) for r in self._read_rows("Contacts")]

    def update_contact(self, contact: Contact) -> None:
        self._write_rows("Contacts", CONTACT_COLUMNS,
                         replace_contact_row(self._read_rows("Contacts"), contact))

    def write_projects(self, projects: list[Project]) -> None:
        self._write_rows("Projects", PROJECT_COLUMNS,
                         merge_project_rows(self._read_rows("Projects"), projects))

    def read_projects(self) -> list[Project]:
        return [row_to_project(r) for r in self._read_rows("Projects")]

    def write_jobs(self, jobs: list[Job]) -> None:
        from . import config

        self._write_rows("Jobs", JOB_COLUMNS,
                         merge_job_rows(self._read_rows("Jobs"), jobs,
                                        floor=config.jobs_min_fit()))

    def read_jobs(self) -> list[Job]:
        return [row_to_job(r) for r in self._read_rows("Jobs")]


def get_tracker() -> Tracker:
    """Return the active tracker per RELAY_TRACKER_BACKEND (default: local xlsx)."""
    from . import config

    backend = config.tracker_backend()
    if backend == "sheets":
        key = config.sheets_workbook_key()
        if not key:
            raise RuntimeError("SHEETS_WORKBOOK_KEY not set (see .env.example)")
        return SheetsTracker(key)
    if backend == "xlsx":
        return LocalXlsxTracker(config.workbook_path())
    raise RuntimeError(f"unknown RELAY_TRACKER_BACKEND: {backend!r} (expected xlsx|sheets)")
