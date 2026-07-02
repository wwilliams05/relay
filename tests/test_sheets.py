"""LocalXlsxTracker round-trips: upserts, human-gate preservation, sort + pruning."""

from __future__ import annotations

import zipfile
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from relay import config
from relay.models import Contact, EmailStatus, Job, Why
from relay.sheets import LocalXlsxTracker, contact_key, job_key


def _check_box(path: Path, sheet: str, key_header: str, key_value: str, bool_header: str) -> None:
    """Simulate the human ticking a checkbox in Excel: set the boolean cell to TRUE."""
    wb = load_workbook(path)
    ws = wb[sheet]
    headers = [c.value for c in ws[1]]
    key_col, bool_col = headers.index(key_header) + 1, headers.index(bool_header) + 1
    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row=row, column=key_col).value) == key_value:
            ws.cell(row=row, column=bool_col, value=True)
    wb.save(path)


def _contact(**overrides) -> Contact:
    base = dict(
        name="Elan Reyes", title="Business Operations Analyst, Starlink", company="SpaceX",
        profile_url="https://www.linkedin.com/in/elan-reyes-example", why=Why.ALUMNI,
        school_match="University of Southern California", email="elan.reyes@spacex.com",
        email_status=EmailStatus.VERIFIED, hook="USC grad who moved from consulting into Starlink ops",
    )
    base.update(overrides)
    return Contact(**base)


def _job(**overrides) -> Job:
    base = dict(
        company="Stripe", title="Product Management Intern", location="San Francisco, CA",
        source="greenhouse", job_url="https://boards.example/stripe/pm-intern",
        date_posted=date(2026, 6, 5), fit_score=80, fit_reason="matches your target role",
    )
    base.update(overrides)
    return Job(**base)


# --- identity keys ---------------------------------------------------------------
def test_job_key_prefers_url() -> None:
    assert job_key(_job()) == "url::https://boards.example/stripe/pm-intern"
    assert job_key(_job(job_url=None)) == "ct::stripe::product management intern"


def test_contact_key_prefers_email() -> None:
    assert contact_key(_contact()) == "email::elan.reyes@spacex.com"
    assert contact_key(_contact(email=None)) == "nc::elan reyes::spacex"


# --- contacts ---------------------------------------------------------------------
def test_contacts_round_trip(tracker: LocalXlsxTracker) -> None:
    original = _contact(
        want_to_message=True, referral_cleared=False, draft_created=True,
        messaged_date=date(2026, 7, 1), responded=True,
        chat_notes="Talked Starlink ops tooling", next_step="Send thank-you note",
    )
    tracker.write_contacts([original])
    (read,) = tracker.read_contacts()
    assert read == original


def test_write_contacts_preserves_human_columns_on_rerun(tracker: LocalXlsxTracker) -> None:
    tracker.write_contacts([_contact()])
    _check_box(tracker.path, "Contacts", "name", "Elan Reyes", "want_to_message")

    # Re-discovery returns the same person (fresher machine data) + a new one.
    tracker.write_contacts([_contact(title="Senior BizOps Analyst"), _contact(
        name="Dana Okoro", email="dana.okoro@spacex.com", why=Why.SIMILAR_ROLE,
        school_match=None)])
    by_name = {c.name: c for c in tracker.read_contacts()}
    assert by_name["Elan Reyes"].want_to_message is True  # human tick survived
    assert by_name["Elan Reyes"].title == "Senior BizOps Analyst"  # machine data refreshed
    assert by_name["Dana Okoro"].want_to_message is False


def test_update_contact_overwrites_single_row(tracker: LocalXlsxTracker) -> None:
    tracker.write_contacts([_contact(), _contact(name="Dana Okoro", email="dana.okoro@spacex.com")])
    changed = _contact(draft_created=True, messaged_date=date(2026, 7, 2))
    tracker.update_contact(changed)
    by_name = {c.name: c for c in tracker.read_contacts()}
    assert by_name["Elan Reyes"].draft_created is True
    assert by_name["Elan Reyes"].messaged_date == date(2026, 7, 2)
    assert len(by_name) == 2


# --- jobs ---------------------------------------------------------------------------
def test_write_jobs_sorts_highest_fit_first(tracker: LocalXlsxTracker) -> None:
    tracker.write_jobs([
        _job(job_url="https://x.example/a", fit_score=30),
        _job(job_url="https://x.example/b", fit_score=90),
        _job(job_url="https://x.example/c", fit_score=50),
    ])
    assert [j.fit_score for j in tracker.read_jobs()] == [90, 50, 30]


def test_write_jobs_prunes_below_floor_unless_pursued(tracker: LocalXlsxTracker) -> None:
    tracker.write_jobs([
        _job(job_url="https://x.example/keep", fit_score=55),
        _job(job_url="https://x.example/noise", fit_score=5),   # below the 20 floor
    ])
    assert [j.fit_score for j in tracker.read_jobs()] == [55]

    # A pursued job survives even when its score sits under the floor.
    tracker.write_jobs([_job(job_url="https://x.example/low", fit_score=10, pursue=True)])
    urls = {j.job_url for j in tracker.read_jobs()}
    assert "https://x.example/low" in urls


def test_write_jobs_preserves_pursue_and_refreshes_score(tracker: LocalXlsxTracker) -> None:
    tracker.write_jobs([_job(fit_score=50)])
    _check_box(tracker.path, "Jobs", "title", "Product Management Intern", "pursue")

    tracker.write_jobs([_job(fit_score=60)])  # re-run rescored the same URL
    (read,) = tracker.read_jobs()
    assert read.pursue is True
    assert read.fit_score == 60


def test_jobs_round_trip_fields(tracker: LocalXlsxTracker) -> None:
    original = _job(pursue=True, status="applied")
    tracker.write_jobs([original])
    (read,) = tracker.read_jobs()
    assert read == original


# --- checkbox injection ----------------------------------------------------------------
def test_checkbox_injection_round_trip(monkeypatch, tracker: LocalXlsxTracker) -> None:
    monkeypatch.setenv("RELAY_XLSX_CHECKBOXES", "1")
    tracker.write_jobs([_job()])
    with zipfile.ZipFile(tracker.path) as z:
        assert "xl/featurePropertyBag/featurePropertyBag.xml" in z.namelist()
    # The injected workbook still reads back cleanly.
    (read,) = tracker.read_jobs()
    assert read.company == "Stripe"
    assert read.pursue is False
    assert config.xlsx_checkboxes() is True
