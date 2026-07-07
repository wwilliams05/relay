"""Shared fixtures. Tests run hermetically: fixture adapters, a temp workbook, and
plain-boolean xlsx (checkbox injection gets its own explicit test)."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from relay.models import Profile


def check_box(path: Path, sheet: str, key_header: str, key_value: str,
              bool_header: str) -> None:
    """Simulate the human ticking a checkbox in Excel: set the boolean cell to TRUE."""
    wb = load_workbook(path)
    ws = wb[sheet]
    headers = [c.value for c in ws[1]]
    key_col, bool_col = headers.index(key_header) + 1, headers.index(bool_header) + 1
    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row=row, column=key_col).value) == key_value:
            ws.cell(row=row, column=bool_col, value=True)
    wb.save(path)


@pytest.fixture(autouse=True)
def hermetic_env(monkeypatch, tmp_path: Path) -> Path:
    """Pin every env-driven mode so tests never touch the network, real workbook,
    or the user's saved profile — regardless of the developer's local .env."""
    monkeypatch.setenv("RELAY_JOBS_MODE", "fixture")
    monkeypatch.setenv("RELAY_APOLLO_MODE", "fixture")
    monkeypatch.setenv("RELAY_GMAIL_MODE", "fixture")
    monkeypatch.setenv("RELAY_TRACKER_BACKEND", "xlsx")
    monkeypatch.setenv("RELAY_WORKBOOK_PATH", str(tmp_path / "relay.xlsx"))
    monkeypatch.setenv("RELAY_PROFILE_PATH", str(tmp_path / "profile.json"))
    monkeypatch.setenv("RELAY_DRAFTS_DIR", str(tmp_path / "drafts"))
    monkeypatch.setenv("RELAY_JOBS_MIN_FIT", "20")
    monkeypatch.setenv("RELAY_XLSX_CHECKBOXES", "0")
    return tmp_path


@pytest.fixture()
def profile() -> Profile:
    """A representative parsed-résumé profile (mirrors the v1 dogfood user)."""
    return Profile(
        name="Weston Williams",
        schools=["University of Southern California", "Washington University in St. Louis"],
        major="Business Administration",
        roles=["Business Operations Intern"],
        skills=["SQL", "Excel"],
        preferred_locations=["Los Angeles", "New York", "Remote"],
        extra_context="Fall 2026 Co-Op, Product Management or BizOps",
    )


@pytest.fixture()
def tracker(hermetic_env):
    from relay import config
    from relay.sheets import LocalXlsxTracker

    return LocalXlsxTracker(config.workbook_path())
